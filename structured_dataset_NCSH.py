import pandas as pd
import numpy as np
from openpyxl.reader.excel import load_workbook

# Load the DataSet file
file_path = "datasets/NCSH_Foundation_Dataset.xlsx"
data_frame = pd.read_excel(file_path, sheet_name="PA Log Sheet")

# Read Excel and count the total number of rows and columns
wb = load_workbook(file_path)
num_columns = wb['PA Log Sheet'].max_column
num_rows = wb['PA Log Sheet'].max_row
print(f"Total number of Columns count is {num_columns}")
print(f"Total number of Rows count is {num_rows}")

# Analyse Data Types within the Data Set
categorical_cols = ['Request_Status', 'Payment_Submitted?', 'Reason___Pending/No', 'Application_Signed?']
numerical_cols = ['Amount', 'Remaining_Balance']

# Standardize the dataset headers by trimming the trailing whitespaces, replacing and Uppercasing the columns name
data_frame.columns = data_frame.columns.str.strip().str.replace(' ', '_').str.replace('-', '_')
print(data_frame.columns.tolist())

# Replace cells having value as "Missing" with NaN
data_frame.replace(['Missing', 'missing'], np.nan, inplace=True)

# Convert and Clean Dates
data_frame['Grant_Req_Date'] = pd.to_datetime(data_frame['Grant_Req_Date'], errors='coerce')

# Clean Categorical Fields
for col in categorical_cols:
    data_frame.replace([''], np.nan, inplace=True)
    data_frame[col] = data_frame[col].astype(str).str.strip().str.lower()

# Clean Numerical Fields
for col in numerical_cols:
    data_frame[col] = pd.to_numeric(data_frame[col], errors='coerce')

# Create a Columns in the Data Frame with the combination of existing columns from the datasets
data_frame['Ready_for_Review'] = (data_frame['Request_Status'] == 'approved') & (data_frame['Application_Signed?'] != 'yes')

# Simulate variable support sent delays (normal distribution)
np.random.seed(42)  # for reproducibility
delays = np.random.normal(loc=7, scale=2, size=len(data_frame)).round().astype(int)
delays = np.clip(delays, 1, None)  # ensure minimum delay is 1 day

data_frame["Days_To_Support"] = delays
data_frame["Support_Sent_Date"] = data_frame["Grant_Req_Date"] + pd.to_timedelta(data_frame["Days_To_Support"], unit="D")

# Export the cleaned dataset in xlsx format
cleaned_file_path = "datasets/NCSH_Foundation_Dataset_Cleaned.xlsx"
data_frame.to_excel(cleaned_file_path, index=False)

# Convert the cleaned dataset to csv format
output_file = "datasets/NCSH_Foundation_Dataset_Cleaned.csv"
data_frame.to_csv(output_file, index=False)